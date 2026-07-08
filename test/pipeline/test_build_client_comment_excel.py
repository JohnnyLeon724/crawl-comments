import json
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PIPELINE_SRC = ROOT / "src" / "pipeline"
if str(PIPELINE_SRC) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SRC))


class BuildClientCommentExcelTest(unittest.TestCase):
    def write_json(self, path, value):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    def write_jsonl(self, path, rows):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows), encoding="utf-8")

    def test_builds_delivery_workbook_with_standard_sheets(self):
        from build_client_comment_excel import build_delivery_workbook

        with TemporaryDirectory() as tmp:
            project_dir = Path(tmp)
            out = project_dir / "delivery.xlsx"
            self.write_json(project_dir / "crawl-tasks.json", {
                "schema_version": "crawl-tasks-v1",
                "tasks": [
                    {
                        "task_id": "task_0001",
                        "phase": "KOL link-0630",
                        "source_excel_row": 2,
                        "source_index": "1",
                        "platform": "douyin",
                        "creator_name": "DJ初仔大朋友",
                        "published_at_text": "6.15",
                        "source_url": "https://v.douyin.com/abc/",
                        "engagement_count": 134000,
                        "expected_comment_count": 2,
                    }
                ],
            })
            self.write_jsonl(project_dir / "all-normalized-comments.jsonl", [
                {
                    "row_key": "r1",
                    "task_id": "task_0001",
                    "phase": "KOL link-0630",
                    "source_excel_row": 2,
                    "source_index": "1",
                    "creator_name": "DJ初仔大朋友",
                    "published_at_text": "6.15",
                    "platform": "douyin",
                    "source_url": "https://www.douyin.com/video/123",
                    "row_type": "level1",
                    "user_name": "用户A",
                    "text": "主评论",
                    "created_at": "2周前",
                    "like_count": 3,
                    "reply_to_user_name": "",
                    "source_engagement_count": 134000,
                    "source_expected_comment_count": 2,
                },
                {
                    "row_key": "r2",
                    "task_id": "task_0001",
                    "phase": "KOL link-0630",
                    "source_excel_row": 2,
                    "source_index": "1",
                    "creator_name": "DJ初仔大朋友",
                    "published_at_text": "6.15",
                    "platform": "douyin",
                    "source_url": "https://www.douyin.com/video/123",
                    "row_type": "level2",
                    "user_name": "用户B",
                    "text": "回复内容",
                    "created_at": "1周前",
                    "like_count": 1,
                    "reply_to_user_name": "用户A",
                    "source_engagement_count": 134000,
                    "source_expected_comment_count": 2,
                },
            ])

            summary = build_delivery_workbook(project_dir, out)

            self.assertEqual(summary["status"], "success")
            self.assertEqual(summary["summary_rows"], 1)
            self.assertEqual(summary["detail_rows"], 2)
            self.assertTrue(out.exists())

            workbook = load_workbook(out)
            self.assertEqual(workbook.sheetnames, ["汇总", "阶段汇总", "评论明细"])
            summary_sheet = workbook["汇总"]
            phase_sheet = workbook["阶段汇总"]
            detail_sheet = workbook["评论明细"]

            self.assertEqual(summary_sheet["A1"].value, "阶段")
            self.assertEqual(summary_sheet["I2"].value, 2)
            self.assertEqual(summary_sheet["J2"].value, 1)
            self.assertEqual(summary_sheet["K2"].value, 1)
            self.assertEqual(summary_sheet["L2"].value, "ok")

            self.assertEqual(phase_sheet["A2"].value, "KOL link-0630")
            self.assertEqual(phase_sheet["B2"].value, "douyin")
            self.assertEqual(phase_sheet["C2"].value, 1)
            self.assertEqual(phase_sheet["D2"].value, 2)

            self.assertEqual(detail_sheet["J2"].value, "第1楼")
            self.assertEqual(detail_sheet["K2"].value, "主评论")
            self.assertEqual(detail_sheet["N2"].value, "用户A")
            self.assertEqual(detail_sheet["O2"].value, "主评论")
            self.assertEqual(detail_sheet["J3"].value, "第1楼-第1条回复")
            self.assertEqual(detail_sheet["K3"].value, "回复")
            self.assertEqual(detail_sheet["P3"].value, "用户A")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
