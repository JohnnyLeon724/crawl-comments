import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = "crawl-tasks-v1"
REQUIRED_COLUMNS = [
    "序号",
    "账户ID",
    "发布平台",
    "发布日期",
    "发布链接",
    "播放量/曝光",
    "互动总量",
    "评论数",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def read_number(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)

    text = normalize_text(value)
    if not text:
        return 0

    unit_multiplier = 1
    if text.endswith("万"):
        unit_multiplier = 10000
        text = text[:-1]
    elif text.endswith("k") or text.endswith("K"):
        unit_multiplier = 1000
        text = text[:-1]

    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned:
        return 0
    return int(float(cleaned) * unit_multiplier)


def extract_first_url(text: Any) -> str:
    value = normalize_text(text)
    match = re.search(r"https?://[^\s，,，。；;）)]+", value)
    if not match:
        return ""
    return match.group(0).rstrip(".,;，。；")


def normalize_platform(value: Any) -> str:
    text = normalize_text(value).lower()
    compact = re.sub(r"\s+", "", text)

    if "抖音" in compact or "douyin" in compact:
        return "douyin"
    if "小红书" in compact or "xiaohongshu" in compact or "xhs" in compact:
        return "xiaohongshu"
    if "b站" in compact or "bilibili" in compact or "哔哩" in compact:
        return "bilibili"
    if "微博" in compact or "weibo" in compact:
        return "weibo"
    return "unknown"


def build_header_map(sheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for cell in sheet[1]:
        name = normalize_text(cell.value)
        if name:
            header_map[name] = cell.column

    missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"客户需求表缺少必需列：{', '.join(missing)}")

    return header_map


def read_cell(row, header_map: dict[str, int], column_name: str) -> Any:
    return row[header_map[column_name] - 1].value


def build_task(row, header_map: dict[str, int], row_number: int, task_number: int, phase: str) -> dict[str, Any]:
    source_url_text = normalize_text(read_cell(row, header_map, "发布链接"))

    return {
        "task_id": f"task_{task_number:04d}",
        "phase": normalize_text(phase),
        "source_excel_row": row_number,
        "source_index": normalize_text(read_cell(row, header_map, "序号")),
        "platform": normalize_platform(read_cell(row, header_map, "发布平台")),
        "creator_name": normalize_text(read_cell(row, header_map, "账户ID")),
        "published_at_text": normalize_text(read_cell(row, header_map, "发布日期")),
        "source_url_text": source_url_text,
        "source_url": extract_first_url(source_url_text),
        "exposure_count": read_number(read_cell(row, header_map, "播放量/曝光")),
        "engagement_count": read_number(read_cell(row, header_map, "互动总量")),
        "expected_comment_count": read_number(read_cell(row, header_map, "评论数")),
        "status": "pending",
    }


def parse_workbook(input_path: str | Path, phase: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header_map = build_header_map(sheet)
    tasks: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2):
        source_url_text = normalize_text(read_cell(row, header_map, "发布链接"))
        creator_name = normalize_text(read_cell(row, header_map, "账户ID"))
        if not source_url_text and not creator_name:
            continue
        tasks.append(build_task(row, header_map, row[0].row, len(tasks) + 1, phase))

    workbook.close()
    return tasks


def build_manifest(tasks: list[dict[str, Any]], out_dir: str | Path) -> dict[str, Any]:
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    output_dir = Path(out_dir)
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "pending",
        "created_at": now,
        "out_dir": str(output_dir),
        "task_count": len(tasks),
        "pending_count": len([task for task in tasks if task.get("status") == "pending"]),
        "tasks": [
            {
                "task_id": str(task["task_id"]),
                "status": str(task.get("status") or "pending"),
                "run_dir": str(output_dir / "runs" / str(task["task_id"])),
                "task_file": str(output_dir / "runs" / str(task["task_id"]) / "task.json"),
            }
            for task in tasks
        ],
        "output_files": {
            "tasks": str(output_dir / "crawl-tasks.json"),
            "manifest": str(output_dir / "run-manifest.json"),
        },
    }


def write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_project_files(tasks: list[dict[str, Any]], out_dir: str | Path) -> dict[str, Any]:
    output_dir = Path(out_dir)
    tasks_payload = {
        "schema_version": SCHEMA_VERSION,
        "task_count": len(tasks),
        "tasks": tasks,
    }
    manifest = build_manifest(tasks, output_dir)

    write_json(output_dir / "crawl-tasks.json", tasks_payload)
    write_json(output_dir / "run-manifest.json", manifest)
    for task in tasks:
        write_json(output_dir / "runs" / str(task["task_id"]) / "task.json", task)

    return {
        "status": "success",
        "task_count": len(tasks),
        "tasks_file": str(output_dir / "crawl-tasks.json"),
        "manifest_file": str(output_dir / "run-manifest.json"),
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse a client requirement workbook into crawl tasks.")
    parser.add_argument("--input", required=True, help="Client requirement .xlsx file.")
    parser.add_argument("--phase", required=True, help="Batch or phase name to attach to generated tasks.")
    parser.add_argument("--out-dir", required=True, help="Project output directory.")
    parser.add_argument("--sheet", default="", help="Optional sheet name. Defaults to the first sheet.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    tasks = parse_workbook(args.input, phase=args.phase, sheet_name=args.sheet or None)
    result = write_project_files(tasks, args.out_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
