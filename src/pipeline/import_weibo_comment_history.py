import argparse
import hashlib
import json
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook

from parse_client_requirements import (
    extract_first_url,
    normalize_text,
    read_number,
    write_json,
    write_project_files,
)


WEIBO_SUMMARY_SHEET = "微博汇总"
WEIBO_DETAIL_SHEET = "评论明细"
SCHEMA_VERSION = "weibo-history-import-v1"
DETAIL_CARRY_FIELDS = (
    "阶段",
    "Excel行",
    "序号",
    "博主昵称",
    "平台",
    "页面链接",
    "微博互动量",
)


def header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        normalize_text(value): index
        for index, value in enumerate(row)
        if normalize_text(value)
    }


def read_value(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def task_key(phase: Any, excel_row: Any, source_index: Any) -> tuple[str, str, str]:
    return (
        normalize_text(phase),
        normalize_text(excel_row),
        normalize_text(source_index),
    )


def build_row_key(
    source_url: str,
    floor_info: str,
    parent_floor: str,
    reply_index: str,
    user_name: str,
    text: str,
    time_location: str,
) -> str:
    source = "|".join([
        source_url,
        floor_info,
        parent_floor,
        reply_index,
        user_name,
        text,
        time_location,
    ])
    return "weibo-history::" + hashlib.sha256(source.encode("utf-8")).hexdigest()


def classify_row_type(comment_type: str) -> str:
    return "level2" if "回复" in normalize_text(comment_type) else "level1"


def parse_time_location(value: Any) -> tuple[str, str]:
    text = normalize_text(value)
    if " / " in text:
        created_at, location = text.split(" / ", 1)
        return normalize_text(created_at), normalize_text(location)
    return text, ""


def extract_weibo_post_id(source_url: str) -> str:
    try:
        parsed = urlparse(source_url)
    except ValueError:
        return ""
    if parsed.netloc.lower() not in {"weibo.com", "www.weibo.com"}:
        return ""
    segments = [segment for segment in parsed.path.split("/") if segment]
    if len(segments) != 2:
        return ""
    author_or_detail, status_token = segments
    if author_or_detail == "detail":
        return f"detail/{status_token}"
    if author_or_detail.isdigit():
        return f"{author_or_detail}/{status_token}"
    return ""


def read_summary_tasks(workbook) -> tuple[list[dict[str, Any]], dict[tuple[str, str, str], dict[str, Any]]]:
    if WEIBO_SUMMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"缺少 sheet：{WEIBO_SUMMARY_SHEET}")

    rows = workbook[WEIBO_SUMMARY_SHEET].iter_rows(values_only=True)
    try:
        headers = header_map(next(rows))
    except StopIteration as error:
        raise ValueError("微博汇总 sheet 为空") from error

    tasks: list[dict[str, Any]] = []
    task_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        phase = normalize_text(read_value(row, headers, "阶段"))
        source_url_text = normalize_text(read_value(row, headers, "页面链接"))
        creator_name = normalize_text(read_value(row, headers, "博主昵称"))
        if not phase and not source_url_text and not creator_name:
            continue

        source_excel_row = read_number(read_value(row, headers, "Excel行"))
        source_index = normalize_text(read_value(row, headers, "序号"))
        source_url = extract_first_url(source_url_text) or source_url_text
        task = {
            "task_id": f"task_{len(tasks) + 1:04d}",
            "phase": phase,
            "source_excel_row": source_excel_row,
            "source_index": source_index,
            "platform": "weibo",
            "creator_name": creator_name,
            "published_at_text": "",
            "source_url_text": source_url_text,
            "source_url": source_url,
            "exposure_count": 0,
            "engagement_count": read_number(read_value(row, headers, "微博互动量")),
            "expected_comment_count": read_number(read_value(row, headers, "微博评论总数")),
            "status": "pending",
            "raw_import_status": normalize_text(read_value(row, headers, "状态")),
        }
        tasks.append(task)
        task_by_key[task_key(phase, source_excel_row, source_index)] = task

    return tasks, task_by_key


def read_detail_comments(
    workbook,
    task_by_key: dict[tuple[str, str, str], dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    if WEIBO_DETAIL_SHEET not in workbook.sheetnames:
        raise ValueError(f"缺少 sheet：{WEIBO_DETAIL_SHEET}")

    rows = workbook[WEIBO_DETAIL_SHEET].iter_rows(values_only=True)
    try:
        headers = header_map(next(rows))
    except StopIteration:
        return [], {
            "level1_count": 0,
            "level2_count": 0,
            "duplicate_row_count": 0,
            "orphan_row_count": 0,
            "missing_source_url_count": 0,
            "missing_text_count": 0,
        }

    carry: dict[str, Any] = {name: "" for name in DETAIL_CARRY_FIELDS}
    root_by_task: dict[str, dict[str, tuple[str, str]]] = {}
    seen_row_keys: set[str] = set()
    comments: list[dict[str, Any]] = []
    counts = {
        "level1_count": 0,
        "level2_count": 0,
        "duplicate_row_count": 0,
        "orphan_row_count": 0,
        "missing_source_url_count": 0,
        "missing_text_count": 0,
    }

    for row_number, row in enumerate(rows, start=2):
        for name in DETAIL_CARRY_FIELDS:
            value = read_value(row, headers, name)
            if normalize_text(value):
                carry[name] = value

        floor_info = normalize_text(read_value(row, headers, "楼层信息"))
        user_name = normalize_text(read_value(row, headers, "评论人"))
        text = normalize_text(read_value(row, headers, "评论内容"))
        if not floor_info and not user_name and not text:
            continue

        key = task_key(carry["阶段"], carry["Excel行"], carry["序号"])
        task = task_by_key.get(key)
        if not task:
            counts["orphan_row_count"] += 1
            continue

        comment_type = normalize_text(read_value(row, headers, "评论类型"))
        parent_floor = normalize_text(read_value(row, headers, "父楼层"))
        reply_index = normalize_text(read_value(row, headers, "回复序号"))
        time_location = normalize_text(read_value(row, headers, "评论时间/地区"))
        created_at, ip_location = parse_time_location(time_location)
        row_type = classify_row_type(comment_type)
        task_roots = root_by_task.setdefault(str(task["task_id"]), {})

        if row_type == "level1":
            root_user_name, root_text = user_name, text
            task_roots[floor_info] = (root_user_name, root_text)
        else:
            root_user_name, root_text = task_roots.get(parent_floor, ("", ""))

        source_url = str(task["source_url"])
        row_key = build_row_key(
            source_url,
            floor_info,
            parent_floor,
            reply_index,
            user_name,
            text,
            time_location,
        )
        if row_key in seen_row_keys:
            counts["duplicate_row_count"] += 1
            continue
        seen_row_keys.add(row_key)

        if not source_url:
            counts["missing_source_url_count"] += 1
        if not text:
            counts["missing_text_count"] += 1

        comment = {
            "row_key": row_key,
            "task_id": task["task_id"],
            "phase": task["phase"],
            "source_excel_row": task["source_excel_row"],
            "source_index": task["source_index"],
            "creator_name": task["creator_name"],
            "published_at_text": task["published_at_text"],
            "source_engagement_count": task["engagement_count"],
            "source_expected_comment_count": task["expected_comment_count"],
            "platform": "weibo",
            "source_url": source_url,
            "post_id": extract_weibo_post_id(source_url),
            "row_type": row_type,
            "comment_id": "",
            "root_comment_id": "",
            "parent_comment_id": "",
            "user_name": user_name,
            "text": text,
            "created_at": created_at,
            "ip_location": ip_location,
            "like_count": read_number(read_value(row, headers, "点赞数")),
            "reply_to_user_name": normalize_text(read_value(row, headers, "回复对象")),
            "root_text": root_text,
            "raw": {
                "importer": "weibo_comment_history",
                "source_sheet": WEIBO_DETAIL_SHEET,
                "source_row": row_number,
                "floor_info": floor_info,
                "parent_floor": parent_floor,
                "reply_index": reply_index,
                "root_user_name": root_user_name,
                "post_text": "",
            },
        }
        comments.append(comment)
        counts[f"{row_type}_count"] += 1

    return comments, counts


def rows_to_jsonl(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    return "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows)


def import_weibo_comment_history(input_path: str | Path, out_dir: str | Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        tasks, task_by_key = read_summary_tasks(workbook)
        comments, counts = read_detail_comments(workbook, task_by_key)
    finally:
        workbook.close()

    output_dir = Path(out_dir)
    project_summary = write_project_files(tasks, output_dir)
    comments_path = output_dir / "all-normalized-comments.jsonl"
    comments_path.write_text(rows_to_jsonl(comments), encoding="utf-8")
    import_summary = {
        "schema_version": SCHEMA_VERSION,
        "input": str(Path(input_path)),
        "task_count": len(tasks),
        "comment_count": len(comments),
        **counts,
    }
    summary_path = output_dir / "history-import-summary.json"
    write_json(summary_path, import_summary)

    return {
        "status": "success",
        "task_count": len(tasks),
        "comment_count": len(comments),
        "tasks_file": project_summary["tasks_file"],
        "manifest_file": project_summary["manifest_file"],
        "comments_file": str(comments_path),
        "summary_file": str(summary_path),
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import a historical Weibo comment workbook into the standard comment pipeline format.")
    parser.add_argument("--input", required=True, help="Historical Weibo .xlsx file.")
    parser.add_argument("--out-dir", required=True, help="Output project directory.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    result = import_weibo_comment_history(args.input, args.out_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
