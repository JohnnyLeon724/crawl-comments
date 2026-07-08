import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_COLUMNS = [
    "阶段",
    "Excel行",
    "序号",
    "平台",
    "博主昵称",
    "发布日期",
    "源表链接",
    "实际打开URL",
    "已抓评论数",
    "主评论数",
    "回复数",
    "状态",
    "备注",
]

PHASE_COLUMNS = [
    "阶段",
    "平台",
    "内容数",
    "已抓评论数",
    "主评论数",
    "回复数",
]

DETAIL_COLUMNS = [
    "阶段",
    "Excel行",
    "序号",
    "博主昵称",
    "平台",
    "发布日期",
    "页面链接",
    "内容互动量",
    "源表评论数",
    "楼层信息",
    "评论类型",
    "父楼层",
    "回复序号",
    "评论人",
    "评论内容",
    "回复对象",
    "点赞数",
    "评论时间/地区",
]


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if stripped:
            rows.append(json.loads(stripped))
    return rows


def read_tasks(project_dir: Path) -> list[dict[str, Any]]:
    payload = read_json(project_dir / "crawl-tasks.json")
    tasks = payload.get("tasks") if isinstance(payload, dict) else []
    return [task for task in tasks if isinstance(task, dict)]


def read_qa_by_task(project_dir: Path) -> dict[str, dict[str, Any]]:
    qa_path = project_dir / "qa-summary.json"
    if not qa_path.exists():
        return {}

    payload = read_json(qa_path)
    tasks = payload.get("tasks") if isinstance(payload, dict) else []
    return {
        str(task.get("task_id") or ""): task
        for task in tasks
        if isinstance(task, dict) and task.get("task_id")
    }


def group_comments_by_task(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        task_id = str(row.get("task_id") or "")
        if task_id:
            grouped[task_id].append(row)
    return grouped


def count_rows(rows: list[dict[str, Any]], row_type: str) -> int:
    return len([row for row in rows if row.get("row_type") == row_type])


def task_status(task: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    expected = int(task.get("expected_comment_count") or 0)
    if not rows:
        return "failed" if expected else "ok"
    if expected and len(rows) / expected < 0.8:
        return "partial"
    return "ok"


def first_actual_url(task: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    for row in rows:
        if row.get("source_url"):
            return str(row["source_url"])
    return str(task.get("source_url") or "")


def build_summary_rows(
    tasks: list[dict[str, Any]],
    comments_by_task: dict[str, list[dict[str, Any]]],
    qa_by_task: dict[str, dict[str, Any]] | None = None,
) -> list[list[Any]]:
    qa_by_task = qa_by_task or {}
    rows = []
    for task in tasks:
        task_id = str(task.get("task_id") or "")
        comments = comments_by_task.get(task_id, [])
        qa = qa_by_task.get(task_id, {})
        rows.append([
            task.get("phase", ""),
            task.get("source_excel_row", ""),
            task.get("source_index", ""),
            task.get("platform", ""),
            task.get("creator_name", ""),
            task.get("published_at_text", ""),
            task.get("source_url", ""),
            first_actual_url(task, comments),
            len(comments),
            count_rows(comments, "level1"),
            count_rows(comments, "level2"),
            qa.get("status") or task_status(task, comments),
            qa.get("notes", ""),
        ])
    return rows


def build_phase_rows(tasks: list[dict[str, Any]], comments_by_task: dict[str, list[dict[str, Any]]]) -> list[list[Any]]:
    aggregates: dict[tuple[str, str], dict[str, int]] = {}

    for task in tasks:
        phase = str(task.get("phase") or "")
        platform = str(task.get("platform") or "")
        key = (phase, platform)
        comments = comments_by_task.get(str(task.get("task_id") or ""), [])
        bucket = aggregates.setdefault(key, {
            "content_count": 0,
            "comment_count": 0,
            "level1_count": 0,
            "level2_count": 0,
        })
        bucket["content_count"] += 1
        bucket["comment_count"] += len(comments)
        bucket["level1_count"] += count_rows(comments, "level1")
        bucket["level2_count"] += count_rows(comments, "level2")

    return [
        [phase, platform, values["content_count"], values["comment_count"], values["level1_count"], values["level2_count"]]
        for (phase, platform), values in aggregates.items()
    ]


def combine_time_location(row: dict[str, Any]) -> str:
    created_at = str(row.get("created_at") or "")
    ip_location = str(row.get("ip_location") or row.get("raw", {}).get("ai_row", {}).get("ip_location") or "")
    if created_at and ip_location:
        return f"{created_at}·{ip_location}"
    return created_at or ip_location


def build_detail_rows(tasks: list[dict[str, Any]], comments_by_task: dict[str, list[dict[str, Any]]]) -> list[list[Any]]:
    detail_rows: list[list[Any]] = []

    for task in tasks:
        comments = comments_by_task.get(str(task.get("task_id") or ""), [])
        floor = 0
        reply = 0
        parent_floor = ""

        for row in comments:
            if row.get("row_type") == "level2":
                reply += 1
                floor_info = f"{parent_floor}-第{reply}条回复" if parent_floor else f"第{reply}条回复"
                comment_type = "回复"
                parent = parent_floor
                reply_index: Any = reply
            else:
                floor += 1
                reply = 0
                parent_floor = f"第{floor}楼"
                floor_info = parent_floor
                comment_type = "主评论"
                parent = ""
                reply_index = ""

            detail_rows.append([
                row.get("phase") or task.get("phase", ""),
                row.get("source_excel_row") or task.get("source_excel_row", ""),
                row.get("source_index") or task.get("source_index", ""),
                row.get("creator_name") or task.get("creator_name", ""),
                row.get("platform") or task.get("platform", ""),
                row.get("published_at_text") or task.get("published_at_text", ""),
                row.get("source_url") or task.get("source_url", ""),
                row.get("source_engagement_count") or task.get("engagement_count", 0),
                row.get("source_expected_comment_count") or task.get("expected_comment_count", 0),
                floor_info,
                comment_type,
                parent,
                reply_index,
                row.get("user_name", ""),
                row.get("text", ""),
                row.get("reply_to_user_name", ""),
                int(row.get("like_count") or 0),
                combine_time_location(row),
            ])

    return detail_rows


def write_sheet(sheet, columns: list[str], rows: list[list[Any]]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 10), 48)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_delivery_workbook(project_dir: str | Path, out: str | Path, template: str | Path | None = None) -> dict[str, Any]:
    root = Path(project_dir)
    output_path = Path(out)
    tasks = read_tasks(root)
    comments = read_jsonl(root / "all-normalized-comments.jsonl")
    comments_by_task = group_comments_by_task(comments)
    qa_by_task = read_qa_by_task(root)

    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "汇总"
    phase_sheet = workbook.create_sheet("阶段汇总")
    detail_sheet = workbook.create_sheet("评论明细")

    summary_rows = build_summary_rows(tasks, comments_by_task, qa_by_task)
    phase_rows = build_phase_rows(tasks, comments_by_task)
    detail_rows = build_detail_rows(tasks, comments_by_task)

    write_sheet(default_sheet, SUMMARY_COLUMNS, summary_rows)
    write_sheet(phase_sheet, PHASE_COLUMNS, phase_rows)
    write_sheet(detail_sheet, DETAIL_COLUMNS, detail_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    return {
        "status": "success",
        "project_dir": str(root),
        "out": str(output_path),
        "summary_rows": len(summary_rows),
        "phase_rows": len(phase_rows),
        "detail_rows": len(detail_rows),
        "template": str(template or ""),
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the standard client delivery comment workbook.")
    parser.add_argument("--project-dir", required=True, help="Project directory containing crawl-tasks.json and all-normalized-comments.jsonl.")
    parser.add_argument("--template", default="", help="Optional reference template path. First version uses standard generated sheets.")
    parser.add_argument("--out", required=True, help="Output .xlsx path.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    summary = build_delivery_workbook(args.project_dir, args.out, args.template or None)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
