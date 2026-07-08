import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from parse_client_requirements import (
    extract_first_url,
    normalize_text,
    read_number,
    write_project_files,
)


BILIBILI_SUMMARY_SHEET = "B站汇总"
BILIBILI_DETAIL_SHEET = "评论明细"


def header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        normalize_text(value): index
        for index, value in enumerate(row)
        if normalize_text(value)
    }


def read_value(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def task_key(phase: Any, excel_row: Any, source_index: Any) -> tuple[str, str, str]:
    return (
        normalize_text(phase),
        normalize_text(excel_row),
        normalize_text(source_index),
    )


def parse_time_location(value: Any) -> tuple[str, str]:
    text = normalize_text(value)
    if " / " in text:
        created_at, location = text.split(" / ", 1)
    elif "/" in text:
        created_at, location = text.split("/", 1)
    elif "·" in text:
        created_at, location = text.split("·", 1)
    else:
        return text, ""
    return normalize_text(created_at), normalize_text(location)


def build_row_key(task_id: str, floor_info: str, user_name: str, text: str) -> str:
    digest = hashlib.sha1(f"{task_id}|{floor_info}|{user_name}|{text}".encode("utf-8")).hexdigest()[:12]
    return f"bilibili::{task_id}::{digest}"


def read_summary_tasks(workbook) -> tuple[list[dict[str, Any]], dict[tuple[str, str, str], dict[str, Any]]]:
    if BILIBILI_SUMMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"缺少 sheet：{BILIBILI_SUMMARY_SHEET}")

    sheet = workbook[BILIBILI_SUMMARY_SHEET]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = header_map(next(rows))
    except StopIteration as error:
        raise ValueError("B站汇总 sheet 为空") from error

    tasks: list[dict[str, Any]] = []
    task_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in rows:
        phase = normalize_text(read_value(row, headers, "阶段"))
        source_url_text = normalize_text(read_value(row, headers, "页面链接"))
        creator_name = normalize_text(read_value(row, headers, "博主昵称"))
        if not phase and not source_url_text and not creator_name:
            continue

        source_excel_row = read_number(read_value(row, headers, "Excel行"))
        source_index = normalize_text(read_value(row, headers, "序号"))
        source_url = extract_first_url(source_url_text) or source_url_text
        task = {
            "task_id": f"task_{len(tasks) + 1:04d}",
            "phase": phase,
            "source_excel_row": source_excel_row,
            "source_index": source_index,
            "platform": "bilibili",
            "creator_name": creator_name,
            "published_at_text": "",
            "source_url_text": source_url_text,
            "source_url": source_url,
            "exposure_count": 0,
            "engagement_count": read_number(read_value(row, headers, "B站互动量")),
            "expected_comment_count": read_number(
                read_value(row, headers, "评论总数")
                or read_value(row, headers, "总评论行数")
            ),
            "status": "pending",
            "raw_import_status": normalize_text(read_value(row, headers, "状态")),
        }
        tasks.append(task)
        task_by_key[task_key(phase, source_excel_row, source_index)] = task

    return tasks, task_by_key


def read_detail_comments(workbook, task_by_key: dict[tuple[str, str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    if BILIBILI_DETAIL_SHEET not in workbook.sheetnames:
        return []

    sheet = workbook[BILIBILI_DETAIL_SHEET]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = header_map(next(rows))
    except StopIteration:
        return []

    comments: list[dict[str, Any]] = []
    carry = {
        "阶段": "",
        "Excel行": "",
        "序号": "",
        "博主昵称": "",
        "平台": "B站",
        "页面链接": "",
        "B站互动量": "",
    }

    for row_number, row in enumerate(rows, start=2):
        for name in carry:
            value = read_value(row, headers, name)
            if normalize_text(value):
                carry[name] = value

        text = normalize_text(read_value(row, headers, "评论内容"))
        user_name = normalize_text(read_value(row, headers, "评论人"))
        floor_info = normalize_text(read_value(row, headers, "楼层信息"))
        if not text and not user_name and not floor_info:
            continue

        key = task_key(carry["阶段"], carry["Excel行"], carry["序号"])
        task = task_by_key.get(key)
        if not task:
            continue

        comment_type = normalize_text(read_value(row, headers, "评论类型"))
        created_at, ip_location = parse_time_location(read_value(row, headers, "评论时间/地区"))
        row_type = "level2" if "回复" in comment_type else "level1"
        comment = {
            "row_key": build_row_key(str(task["task_id"]), floor_info, user_name, text),
            "task_id": task["task_id"],
            "phase": task["phase"],
            "source_excel_row": task["source_excel_row"],
            "source_index": task["source_index"],
            "creator_name": task["creator_name"],
            "published_at_text": task.get("published_at_text", ""),
            "platform": "bilibili",
            "source_url": task["source_url"],
            "row_type": row_type,
            "user_name": user_name,
            "text": text,
            "created_at": created_at,
            "ip_location": ip_location,
            "like_count": read_number(read_value(row, headers, "点赞数")),
            "reply_to_user_name": normalize_text(read_value(row, headers, "回复对象")),
            "source_engagement_count": task["engagement_count"],
            "source_expected_comment_count": task["expected_comment_count"],
            "raw": {
                "importer": "bilibili_delivery",
                "source_sheet": BILIBILI_DETAIL_SHEET,
                "source_row": row_number,
                "floor_info": floor_info,
                "parent_floor": normalize_text(read_value(row, headers, "父楼层")),
                "reply_index": normalize_text(read_value(row, headers, "回复序号")),
            },
        }
        comments.append(comment)

    return comments


def rows_to_jsonl(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    return "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows)


def import_bilibili_delivery(input_path: str | Path, out_dir: str | Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        tasks, task_by_key = read_summary_tasks(workbook)
        comments = read_detail_comments(workbook, task_by_key)
    finally:
        workbook.close()

    output_dir = Path(out_dir)
    project_summary = write_project_files(tasks, output_dir)
    comments_path = output_dir / "all-normalized-comments.jsonl"
    comments_path.write_text(rows_to_jsonl(comments), encoding="utf-8")

    return {
        "status": "success",
        "task_count": len(tasks),
        "comment_count": len(comments),
        "tasks_file": project_summary["tasks_file"],
        "comments_file": str(comments_path),
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import an existing Bilibili delivery workbook into the standard comment pipeline project format.")
    parser.add_argument("--input", required=True, help="Bilibili delivery .xlsx file.")
    parser.add_argument("--out-dir", required=True, help="Output project directory.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    result = import_bilibili_delivery(args.input, args.out_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
